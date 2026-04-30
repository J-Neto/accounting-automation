import logging
from datetime import datetime, date
from openpyxl import load_workbook
from src.transactions import Transaction

logger = logging.getLogger(__name__)


def read_spreadsheet(file_path: str) -> list[Transaction]:
    """
    Reads the transactions spreadsheet and returns a list of Transaction objects.

    Validates each row individually — invalid rows are skipped
    and logged without interrupting the remaining rows.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        List of valid Transaction objects ready for insertion.

    Raises:
        FileNotFoundError: If the file does not exist at the given path.
        KeyError:          If the 'Lançamentos' sheet does not exist in the spreadsheet.
    """
    logger.info("Reading spreadsheet: '%s'", file_path)

    try:
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        logger.critical("File not found: '%s'", file_path)
        raise

    try:
        sheet = workbook['Lançamentos']
    except KeyError:
        logger.critical("Sheet 'Lançamentos' not found in '%s'", file_path)
        raise

    transactions = []
    errors = 0

    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True),
        start=2
    ):
        if not any(row):
            logger.debug("Row %d is empty, skipping.", row_number)
            continue

        try:
            transaction = _build_transaction(row, row_number)
            transactions.append(transaction)
            logger.debug("Row %d read successfully: '%s'", row_number, transaction.description)

        except (ValueError, TypeError) as err:
            errors += 1
            logger.warning("Row %d skipped due to invalid data: %s", row_number, err)

    logger.info(
        "Reading complete: %d valid transactions, %d row(s) skipped.",
        len(transactions), errors
    )

    return transactions


def _build_transaction(row: tuple, row_number: int) -> Transaction:
    """
    Extracts and converts the fields from a spreadsheet row into a Transaction.

    Internal function — the _ prefix indicates it should not be called
    directly from outside this module.

    Args:
        row:        Tuple with the row values (values_only=True).
        row_number: Row number in Excel, used in error messages.

    Returns:
        Validated Transaction object.

    Raises:
        ValueError: If any required field is missing or invalid.
        TypeError:  If the data type cannot be converted.
    """
    description, amount, date_, type_, category, status = _unpack_row(row, row_number)

    converted_amount = _convert_amount(amount, row_number)
    converted_date = _convert_date(date_, row_number)

    return Transaction(
        description=description,
        amount=converted_amount,
        date=converted_date,
        type=type_,
        category=category,
        status=status
    )


def _unpack_row(row: tuple, row_number: int) -> tuple:
    """
    Ensures the row has exactly 6 fields and no required field is None.

    Raises:
        ValueError: If the row has fewer than 6 columns or has empty fields.
    """
    EXPECTED_FIELDS = 6

    if len(row) < EXPECTED_FIELDS:
        raise ValueError(
            f"Row {row_number} has {len(row)} column(s), expected {EXPECTED_FIELDS}."
        )

    description, amount, date_, type_, category, status = row[:EXPECTED_FIELDS]

    required_fields = {
        'description': description,
        'amount':      amount,
        'date':        date_,
        'type':        type_,
        'category':    category,
        'status':      status
    }

    for field_name, content in required_fields.items():
        if content is None or str(content).strip() == '':
            raise ValueError(f"Row {row_number}: field '{field_name}' is empty.")

    return description, amount, date_, type_, category, status


def _convert_amount(amount, row_number: int) -> str:
    """
    Converts the amount to a string with decimal point.

    Accepts: int (1500), float (1500.5), string ("1500" or "1500,50")
    Returns: "1500.00", "1500.50"

    Raises:
        ValueError: If the amount cannot be converted to a number.
    """
    try:
        if isinstance(amount, str):
            amount = amount.replace('.', '').replace(',', '.')

        amount_float = float(amount)

        if amount_float <= 0:
            raise ValueError(
                f"Row {row_number}: amount must be positive, received: {amount_float}"
            )

        return f"{amount_float:.2f}"

    except (TypeError, ValueError) as err:
        raise ValueError(
            f"Row {row_number}: invalid amount '{amount}'. Details: {err}"
        ) from err


def _convert_date(date_value, row_number: int) -> str:
    """
    Converts the date to a string in DD/MM/YYYY format.

    Accepts: datetime object from openpyxl or formatted string.
    Returns: "10/01/2024"

    Raises:
        ValueError: If the date cannot be converted.
    """
    try:
        if isinstance(date_value, (datetime, date)):
            return date_value.strftime('%d/%m/%Y')

        if isinstance(date_value, str):
            for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                try:
                    return datetime.strptime(date_value.strip(), fmt).strftime('%d/%m/%Y')
                except ValueError:
                    continue

        raise ValueError(f"Unrecognized date format: '{date_value}'")

    except Exception as err:
        raise ValueError(
            f"Row {row_number}: invalid date '{date_value}'. Details: {err}"
        ) from err